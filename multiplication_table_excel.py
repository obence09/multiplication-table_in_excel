import openpyxl, pyinputplus
#font importálása
from openpyxl.styles import Font

#Creating Excel Workbook
wb = openpyxl.Workbook()

#Renaming the active sheet
active_sheet = wb.active
active_sheet.title = 'Multiplication_Table'

input_number = pyinputplus.inputInt('Please enter an integer to create the multipication table:\n')

#Creating font object
bold12font = Font(size=12, bold=True) 

row_counter = 1
column_counter = 1

#Putting numbers into rows
while row_counter <= input_number:
    active_sheet.cell(row=row_counter+1, column=1).value = row_counter
    #font méret megadása
    active_sheet.cell(row=row_counter+1, column=1).font = bold12font
    row_counter = row_counter + 1

#Putting the numbers into columns
while column_counter <= input_number:
    active_sheet.cell(row=1, column=column_counter+1).value = column_counter
    #font méret megadása
    active_sheet.cell(row=1, column=column_counter+1).font = bold12font
    column_counter = column_counter + 1

#Resetting the loop counters
row_counter = 1
column_counter = 1

#Creating the calculated multiplication table
while column_counter <= input_number:
    while row_counter <= input_number:
        active_sheet.cell(row= row_counter +1, column= column_counter+1).value = active_sheet.cell(row= row_counter +1, column=1).value * active_sheet.cell(row= 1, column= column_counter+1).value
        row_counter = row_counter + 1
    row_counter = 1
    column_counter = column_counter+1

#Saving the excel file - provide your own directory
wb.save('C:\\Users\\CFY\\Desktop\\Python\\Projects\\MultiplicationTable.xlsx')
